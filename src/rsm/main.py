"""Entry points after CLI capture."""
import logging
import os
import re
from datetime import datetime
from typing import Any

import rsm.excel as rsm_excel
import yaml
from openpyxl.worksheet.worksheet import Worksheet
from rsm.utils import fmt_date
from rsm.utils import log_setup
from rsm.utils import parser_cfg


def template_config(loglevel):
    """Dump default yaml config file to stdout.

    Args:
        loglevel (str): see rsm.constants.LOG_LEVELS
    """
    log_setup(loglevel)
    print(parser_cfg(None))


def validate(input, config, loglevel) -> tuple[dict[str, Any], Worksheet, datetime, dict[str, int]]:
    """Validate input xlsx against config.

    Args:
        input    (str) : Path to Input xlsx file
        config   (str) : Path to config (None accepted)
        loglevel (str) : see rsm.constants.LOG_LEVELS

    Returns:
        Config loaded into dict; Worksheet to process; Date to record data against; Column mappings kvp.
    """
    log_setup(loglevel)
    input = os.path.abspath(input)
    if config:
        config = os.path.abspath(config)
        logging.debug(f"Reading yaml file: {config}")
    str_config = parser_cfg(config)
    logging.debug(f"Loading yaml to object")
    yml_conf = yaml.safe_load(str_config)
    logging.debug(f"Validating xlsx")
    ws, monthly_dt, col_map = rsm_excel.validate_xlsx(input, yml_conf)
    logging.info("Files validated successfully")
    return (yml_conf, ws, monthly_dt, col_map)


def convert(input: str, config: str, output: str, loglevel: str) -> tuple[dict[str, Any], list[list[Any]]]:
    """Convert xlsx to csv format as described by config.

    Args:
        input: Path to Input xlsx file
        config: Path to config (None accepted)
        output: Path to output csv (None accepted)
        loglevel: see rsm.constants.LOG_LEVELS

    Returns:
        Config loaded into dict; Data rows.
    """
    yml_conf, ws, monthly_dt, col_map = validate(input, config, loglevel)

    if output is None:
        output = re.sub(".xlsx$", ".csv", input)
    else:
        output = os.path.abspath(output)

    output_order = yml_conf["output_order"]
    data_store = []  # datastore has no header
    # -sig is to force the BOM (Byte Order Marker) at the start of the file
    with open(output, mode="w", encoding="utf-8-sig") as ofh:
        # header
        print(",".join(output_order), file=ofh)
        for r in range(3, ws.max_row + 1):
            if ws.cell(r, 1).value == "TOTALS":
                break
            row = [fmt_date(yml_conf["monthly_observation_sets"]["transform_csv"], monthly_dt)]
            for col in output_order:
                if col == "activity_month":
                    continue
                val = ws.cell(r, col_map[col]).value
                if val is None:
                    val = ""
                row.append(str(val))
            print(",".join(row), file=ofh)
            # now output update the date format so we don't need to run an update on the DB
            row[0] = fmt_date(yml_conf["monthly_observation_sets"]["transform_sql"], monthly_dt)
            # and add to the data store
            data_store.append(row)
    logging.info(f"Converted file written to: {output}")

    return (yml_conf, data_store)


def convert_to_db(input: str, config: str, output: str, loglevel: str):
    """Convert xlsx to csv format and upload to database.

    Args:
        input: Path to Input xlsx file
        config: Path to config (None accepted)
        output: Path to output csv (None accepted)
        loglevel: see rsm.constants.LOG_LEVELS
    """
    yml_conf, db_rows = convert(input, config, output, loglevel)
    import rsm.db as rsm_db

    rsm_db.upload(yml_conf, db_rows)
