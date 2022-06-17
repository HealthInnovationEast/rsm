"""All functions specific to parsing of xlsx."""
import logging
import sys
from datetime import datetime
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet


def merged_range_end(ws: Worksheet, row: int, col_start: int, col_max: int) -> int:
    """Identify the end of a column merged cell.

    Args:
        ws (Worksheet): The worksheet in use.
        row (int): Row to process.
        col_start (int): Column to start search.
        col_max (int): Last column of worksheet (1-based)

    Returns:
        int: The last column index of the merge (1-based)
    """
    r_end = 0
    for i in range(col_start, col_max):
        c = ws.cell(row, i)
        c_t = type(c)
        if c_t == Cell:
            break
        if c_t == MergedCell:
            r_end = i

    return r_end


def ranges_in_row(ws: Worksheet, row: int, col_max: int) -> list[tuple[int, int]]:
    """Identify all merged cell ranges in row.

    Args:
        ws (Worksheet): The worksheet in use.
        row (int): Row to process.
        col_max (int): Last column of worksheet (1-based)

    Returns:
         list[Tuple[int, int]]: start/stop positions (1-based)
    """
    ranges = []
    for i in range(1, col_max):
        c = ws.cell(row, i)
        if type(c) is MergedCell:
            continue
        r_start = i
        r_end = merged_range_end(ws, row, i + 1, col_max)
        if r_end > 0:
            ranges.append((r_start, r_end))
    return ranges


def core_cols(conf: dict[str, Any], ws: Worksheet, core_stop: int) -> dict[str, int]:
    """Find core fields between first colunm and first merged range.

    Args:
        conf (Dict[str,Any]): Config loaded into dict.
        ws (Worksheet): The worksheet in use.
        core_stop (int): Position of the first merged range (1-based)

    Returns:
        Dict[str, int]: Field name to column position kvp.
    """
    row = 2
    col_map = {}
    core_set = conf["core_fields"]
    core_keys = core_set.keys()
    captured = 0
    for i in range(1, core_stop + 1):
        val = ws.cell(row, i).value.replace("\n", "").replace("\r", "")
        if val in core_keys:
            col_map[core_set[val]] = i
            captured += 1
    if captured != len(core_keys):
        logging.error(f"Expecting {len(core_keys)} core fields but only got got {captured}")
        logging.error(f" -> Expected: {', '.join(core_keys)}")
        sys.exit(1)
    return col_map


def monthly_cols(
    conf: dict[str, Any], ws: Worksheet, monthly_range: tuple[int, int], col_map: dict[str, int]
) -> dict[str, int]:
    """Extract the appropriate monthly column positions.

    Args:
        conf (Dict[str,Any]): Config loaded into dict.
        ws (Worksheet): The worksheet in use.
        monthly_range (Tuple[int, int]): Column range to search
        col_map (Dict[str, int]): Existing kvp to extend

    Returns:
        Dict[str, int]: Field name to column position kvp.
    """
    row = 2
    monthly_set = conf["monthly_observations"]
    monthly_keys = monthly_set.keys()
    captured = 0
    for i in range(monthly_range[0], monthly_range[1] + 1):
        val = ws.cell(row, i).value.replace("\n", "").replace("\r", "")
        if val in monthly_keys:
            col_map[monthly_set[val]] = i
            captured += 1
    if captured != len(monthly_keys):
        logging.error(f"Expecting {len(monthly_keys)} monthly observation fields but only got got {captured}")
        logging.error(f" -> Expected: {', '.join(monthly_keys)}")
        sys.exit(1)
    return col_map


def validate_xlsx(xlsx: str, conf: dict[str, Any]) -> tuple[Worksheet, datetime, dict[str, int]]:
    """Check xlsx conforms to specification of config file.

    Args:
        xlsx (str): Excel file to load.
        conf (Dict[str,Any]): Config loaded into dict.

    Returns:
        A tuple containing -
            Worksheet: Worksheet to be used
            datetime: The date field extracted based on config
            dict[str, int]: Column name to position kvps.
    """
    wb = load_workbook(filename=xlsx)

    ws = wb.active

    col_max = ws.max_column + 1
    # scan first row to find col-spans, add to list
    row = 1
    ranges = ranges_in_row(ws, row, col_max)
    cfg_month_sets = conf["monthly_observation_sets"]
    exp_ranges = cfg_month_sets["spans_expected"]
    if len(ranges) != exp_ranges:
        logging.error(f"Expecting {exp_ranges} date spans in row 1, got {len(ranges)}")
        logging.error(f"Edit config changing 'monthly_observation_sets.spans_expected' if this has changed.")
        sys.exit(1)

    # needs to be applied to each row_object
    monthly_date = datetime.strptime(
        ws.cell(1, ranges[cfg_month_sets["span_used_idx"]][0]).value, cfg_month_sets["date_parse"]
    )

    # core columns stop at first range:
    core_stop = ranges[0][0] - 1
    if core_stop < len(conf["core_fields"]):
        logging.error(f"Insifficient column before first date range to support:")
        logging.error(f" -> {', '.join(conf['core_fields'].keys())}")

    # generate the column mappings for each field
    col_map = core_cols(conf, ws, core_stop)
    col_map = monthly_cols(conf, ws, ranges[cfg_month_sets["span_used_idx"]], col_map)
    return (ws, monthly_date, col_map)
